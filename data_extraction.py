from openpyxl import load_workbook
import mail
import os

class DataExtraction():
    def __init__(self, *args):
        pass

    def extract_data_from_excel(self, excel_file = "./mockup_data.xlsx"):
        # download the file from gmail:
        mail.download_file()

        # Load the entire workbook.
        data_excel = load_workbook(excel_file)

        # List all the sheets in the file.
        print("Found the following worksheets:")
        for sheetname in data_excel.sheetnames:
            print(sheetname)
            
            worksheet = data_excel[sheetname]
            all_rows = list(worksheet.rows)
            row_nums = len(all_rows)

            print(f"Found {row_nums} rows of data.")
            
            product_list = []
            for row in all_rows[1:row_nums]:
                _id = row[0].value
                business_partner = row[1].value
                product = row[2].value
                amount = row[3].value
                information_dict = {
                    "id": _id,
                    "company": business_partner,
                    "product": product,
                    "quantity": amount
                }
                product_list.append(information_dict)
        return product_list

#os independent file path
data_file = os.path.join(".", "mockup_data.xlsx")
test = DataExtraction()
print(test.extract_data_from_excel(data_file))
